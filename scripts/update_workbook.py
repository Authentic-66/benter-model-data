"""
update_workbook.py
Reads handicap-logs, results-logs, and roi-logs files and writes sheets
to the model workbook.

  - Picks / ROI:   most-recent file per track (upsert)
  - Results:       ALL files per track; skips sheets that already exist

Sheet naming: "{TRACK} {Month D YYYY} Picks/Results/ROI"
  e.g. "FP May 12 2026 Picks", "CT May 28 2026 Results"

Usage: python update_workbook.py
"""

import re
import sys
from datetime import datetime
from pathlib import Path

try:
    sys.stdout.reconfigure(encoding="utf-8")
except AttributeError:
    pass

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("openpyxl not installed.  Run:  pip install openpyxl")
    sys.exit(1)

# ── Paths ──────────────────────────────────────────────────────────────────────

SCRIPTS       = Path(__file__).parent
BASE          = SCRIPTS.parent
WORKBOOK_PATH = BASE / "model-workbook-versions" / "benter_model_master.xlsx"
HANDICAP_DIR  = SCRIPTS / "handicap-logs"
RESULTS_DIR   = SCRIPTS / "results-logs"
ROI_DIR       = SCRIPTS / "roi-logs"

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF")

# ── Utilities ──────────────────────────────────────────────────────────────────

def fmt_date(date_str):
    """YYYYMMDD → 'May 12 2026'"""
    d = datetime.strptime(date_str, "%Y%m%d")
    return f"{d.strftime('%b')} {d.day} {d.year}"

def sheet_name(track, date_str, suffix):
    return f"{track} {fmt_date(date_str)} {suffix}"

def latest_by_track(folder, prefix):
    """
    Find the most-recent file per track matching PREFIX_TRACK_YYYYMMDD.txt.
    Returns {track: (Path, date_str)}.
    """
    result = {}
    for f in folder.glob(f"{prefix}_*.txt"):
        m = re.match(rf"^{prefix}_([A-Z]+)_(\d{{8}})\.txt$", f.name)
        if m:
            track, date = m.group(1), m.group(2)
            if track == "ALL":
                continue
            if track not in result or date > result[track][1]:
                result[track] = (f, date)
    return result

def all_files_for_prefix(folder, prefix):
    """
    Return every file matching PREFIX_TRACK_YYYYMMDD.txt as a sorted list of
    (track, Path, date_str) tuples.  Used for results so all race days are written.
    """
    glob_hits = list(folder.glob(f"{prefix}_*.txt"))
    print(f"  [debug] glob '{prefix}_*.txt' in {folder.name}/ → {len(glob_hits)} files")

    no_match, skip_all, skip_date, found = 0, 0, 0, []
    pattern = re.compile(rf"^{prefix}_([A-Z]+)_(\d{{8}})\.txt$")
    for f in glob_hits:
        m = pattern.match(f.name)
        if not m:
            print(f"  [debug]   regex no-match: {f.name}")
            no_match += 1
            continue
        track, date = m.group(1), m.group(2)
        if track == "ALL":
            skip_all += 1
            continue
        if date < "20260101":
            skip_date += 1
            continue
        found.append((track, f, date))

    print(f"  [debug] kept={len(found)}  no-match={no_match}  "
          f"skip_ALL={skip_all}  skip_pre2026={skip_date}")
    if glob_hits and no_match == len(glob_hits):
        print(f"  [debug] sample filenames: "
              + ", ".join(f.name for f in glob_hits[:5]))
    return sorted(found)

def upsert_sheet(wb, name):
    """Delete sheet if it already exists, then create fresh."""
    if name in wb.sheetnames:
        del wb[name]
    return wb.create_sheet(name)

def write_headers(ws, headers):
    ws.append(headers)
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"

def auto_width(ws):
    for col in ws.columns:
        width = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(width + 2, 45)

# ── Parsers ────────────────────────────────────────────────────────────────────

def parse_picks(text):
    """picks_*.txt / HANDICAP_*.txt → list of row dicts."""
    picks = []
    for line in text.splitlines():
        line = line.strip()
        if not line or line.startswith("#"):
            continue
        parts = line.split()
        if len(parts) >= 4:
            picks.append({
                "race":   parts[1],
                "horse":  parts[2],
                "signal": parts[3],
                "bets":   parts[4] if len(parts) >= 5 else "WPS",
            })
    return picks


def parse_results_log(text):
    """process_results.py console output → list of race dicts."""
    races = []
    current = None
    in_finishers = False

    for line in text.splitlines():
        # Race header: "  RACE 1  —  CLM  |  6F  |  Dirt  |  $20,000"
        m = re.match(r"^\s+RACE\s+(\d+)\s+[—\-]+\s*(.*)", line)
        if m:
            in_finishers = False
            parts = [p.strip() for p in m.group(2).split("|")]
            current = {
                "race":       m.group(1),
                "conditions": parts[0] if parts else "",
                "distance":   parts[1] if len(parts) > 1 else "",
                "surface":    parts[2] if len(parts) > 2 else "",
                "purse":      parts[3] if len(parts) > 3 else "",
                "finishers":  [],
                "time": "", "trainer": "", "sire": "",
                "win": "", "place": "", "show": "",
            }
            races.append(current)
            continue

        if current is None:
            continue

        # Finisher table header
        if re.search(r"\bPos\b.*\bPP\b.*\bHorse\b", line):
            in_finishers = True
            continue

        if in_finishers:
            stripped = line.strip()
            if not stripped:
                in_finishers = False
                continue
            tokens = stripped.split(None, 2)
            if len(tokens) >= 3 and tokens[0].isdigit() and tokens[1].isdigit():
                pair = tokens[2].rsplit(None, 1)
                if len(pair) == 2 and re.match(r"^[\d.?]+$", pair[1]):
                    current["finishers"].append({
                        "pos": tokens[0], "pp": tokens[1],
                        "horse": pair[0].strip(), "odds": pair[1],
                    })

        # Time / Trainer / Sire
        for field, key in [("Time", "time"), ("Trainer", "trainer"), ("Sire", "sire")]:
            m = re.match(rf"^\s+{field}\s*:\s*(.+)", line)
            if m:
                current[key] = m.group(1).strip()

        # Payouts: "  WIN   $5.00   PLC   $3.20   SHW   $2.40"
        for label, key in [("WIN", "win"), ("PLC", "place"), ("SHW", "show")]:
            m = re.search(rf"{label}\s+(\$[\d.]+)", line)
            if m:
                current[key] = m.group(1)

    return races


def parse_roi_log(text, target_track=None):
    """
    roi_tracker.py console output → list of row dicts.
    If target_track is set, only return rows for that track code.
    """
    rows = []
    current_track = None

    for line in text.splitlines():
        # Track section header: "  ── FAIRMOUNT PARK (FP) ──..."
        m = re.search(r"──\s+.+?\s+\(([A-Z]+)\)", line)
        if m:
            current_track = m.group(1)
            continue

        if target_track and current_track != target_track:
            continue

        # Data row: "  R1    MarketWatch      TRAINER   WPS   1  $6.00  $24.50  +$18.50"
        m = re.match(
            r"\s+R(\d+)\s+(\S+)\s+(\S+)\s+([WPS]+)\s+([?\d]+)"
            r"\s+\$(\d+\.\d+)\s+\$(\d+\.\d+)\s+([+\-]\$[\d.]+)(.*)",
            line,
        )
        if m:
            note = m.group(9).strip().strip("[]")
            rows.append({
                "track":    current_track or "",
                "race":     m.group(1),
                "horse":    m.group(2),
                "signal":   m.group(3),
                "bets":     m.group(4),
                "finish":   m.group(5),
                "invested": m.group(6),
                "returned": m.group(7),
                "pl":       m.group(8),
                "status":   note,
            })

    return rows

# ── Sheet writers ──────────────────────────────────────────────────────────────

def write_picks_sheet(wb, name, picks):
    ws = upsert_sheet(wb, name)
    write_headers(ws, ["Race", "Horse", "Signal", "Bets"])
    for p in picks:
        ws.append([p["race"], p["horse"], p["signal"], p["bets"]])
    auto_width(ws)


def write_results_sheet(wb, name, races):
    ws = upsert_sheet(wb, name)
    write_headers(ws, [
        "Race", "Conditions", "Dist", "Surface", "Purse",
        "Pos", "PP", "Horse", "Odds",
        "Time", "Winner Trainer", "Winner Sire",
        "Win $", "Place $", "Show $",
    ])
    for r in races:
        base      = [r["race"], r["conditions"], r["distance"], r["surface"], r["purse"]]
        race_info = [r["time"], r["trainer"], r["sire"], r["win"], r["place"], r["show"]]
        if r["finishers"]:
            for f in r["finishers"]:
                ws.append(base + [f["pos"], f["pp"], f["horse"], f["odds"]] + race_info)
        else:
            ws.append(base + ["", "", "", ""] + race_info)
    auto_width(ws)


def write_roi_sheet(wb, name, rows):
    ws = upsert_sheet(wb, name)
    write_headers(ws, [
        "Race", "Horse", "Signal", "Bets",
        "Finish", "Invested $", "Returned $", "P/L", "Status",
    ])
    total_inv = total_ret = 0.0
    for r in rows:
        ws.append([
            r["race"], r["horse"], r["signal"], r["bets"],
            r["finish"], r["invested"], r["returned"], r["pl"], r["status"],
        ])
        try:
            total_inv += float(r["invested"])
            total_ret += float(r["returned"])
        except ValueError:
            pass

    if rows:
        pl = total_ret - total_inv
        pl_str  = f"+${pl:.2f}"  if pl  >= 0 else f"-${abs(pl):.2f}"
        roi_pct = (pl / total_inv * 100) if total_inv else 0.0
        roi_str = f"+{roi_pct:.1f}%" if roi_pct >= 0 else f"{roi_pct:.1f}%"
        ws.append([
            "TOTAL", "", "", "", "",
            f"${total_inv:.2f}", f"${total_ret:.2f}", pl_str, f"ROI: {roi_str}",
        ])
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True)

    auto_width(ws)

# ── ROI source discovery ───────────────────────────────────────────────────────

def collect_roi_sources(roi_dir):
    """
    Returns {track: (path, date_str, target_track_or_None)}.
    Prefers individual track files (ROI_CT_*); falls back to ROI_ALL sections.
    """
    sources = {}

    # Pass 1: individual track files
    for f in roi_dir.glob("ROI_*.txt"):
        m = re.match(r"^ROI_([A-Z]+)_(\d{8})\.txt$", f.name)
        if not m:
            continue
        track, date = m.group(1), m.group(2)
        if track == "ALL":
            continue
        if track not in sources or date > sources[track][1]:
            sources[track] = (f, date, None)

    # Pass 2: ROI_ALL files — fill gaps for tracks not covered above
    for f in sorted(roi_dir.glob("ROI_ALL_*.txt"), key=lambda x: x.name, reverse=True):
        m = re.match(r"^ROI_ALL_(\d{8})\.txt$", f.name)
        if not m:
            continue
        date = m.group(1)
        text = f.read_text(encoding="utf-8", errors="replace")
        for tm in re.finditer(r"──\s+.+?\s+\(([A-Z]+)\)", text):
            track = tm.group(1)
            if track not in sources or date > sources[track][1]:
                sources[track] = (f, date, track)

    return sources

# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    if not WORKBOOK_PATH.exists():
        print(f"ERROR: workbook not found:\n  {WORKBOOK_PATH}")
        sys.exit(1)

    print(f"Loading {WORKBOOK_PATH.name} ...")
    wb = openpyxl.load_workbook(WORKBOOK_PATH)
    added = []

    # ── Picks (handicap-logs) ──────────────────────────────────────────────
    if HANDICAP_DIR.exists():
        track_files = latest_by_track(HANDICAP_DIR, "HANDICAP")
        if not track_files:
            print("  handicap-logs/ is empty — run parse_ct.bat / parse_fp.bat first")
        for track, (path, date_str) in sorted(track_files.items()):
            name  = sheet_name(track, date_str, "Picks")
            picks = parse_picks(path.read_text(encoding="utf-8", errors="replace"))
            if picks:
                print(f"  + {name}  ({len(picks)} picks)")
                write_picks_sheet(wb, name, picks)
                added.append(name)
            else:
                print(f"  - {path.name}: no picks found")
    else:
        print("  handicap-logs/ not found — run parse_ct.bat / parse_fp.bat first")

    # ── Results (results-logs) — all files, skip existing sheets ─────────
    if RESULTS_DIR.exists():
        print(f"\nResults dir: {RESULTS_DIR}")
        all_results = all_files_for_prefix(RESULTS_DIR, "RESULTS")
        if not all_results:
            print("  results-logs/ is empty — run process_ct.bat / process_fp.bat first")
        else:
            print(f"  Processing {len(all_results)} results files ...")
        for track, path, date_str in all_results:
            name = sheet_name(track, date_str, "Results")
            if name in wb.sheetnames:
                print(f"  = {name}  (already exists, skipping)")
                continue
            print(f"  ? {path.name} → '{name}' ...", end=" ", flush=True)
            try:
                races = parse_results_log(
                    path.read_text(encoding="utf-8", errors="replace")
                )
                if races:
                    print(f"{len(races)} races")
                    write_results_sheet(wb, name, races)
                    added.append(name)
                else:
                    print("no race data found")
            except Exception as exc:
                print(f"ERROR: {exc}")
    else:
        print("  results-logs/ not found — run process_ct.bat / process_fp.bat first")

    # ── ROI (roi-logs) ────────────────────────────────────────────────────
    if ROI_DIR.exists():
        roi_sources = collect_roi_sources(ROI_DIR)
        if not roi_sources:
            print("  roi-logs/ is empty — run roi_ct.bat / roi_fp.bat first")
        for track, (path, date_str, target) in sorted(roi_sources.items()):
            name = sheet_name(track, date_str, "ROI")
            text = path.read_text(encoding="utf-8", errors="replace")
            rows = parse_roi_log(text, target_track=target or track)
            if rows:
                print(f"  + {name}  ({len(rows)} picks)")
                write_roi_sheet(wb, name, rows)
                added.append(name)
            else:
                print(f"  - {path.name}: no ROI rows for {track}")
    else:
        print("  roi-logs/ not found — run roi_ct.bat / roi_fp.bat first")

    if not added:
        print("\nNo sheets added — run the parse / process / roi batch files first.")
        return

    wb.save(WORKBOOK_PATH)
    print(f"\nSaved {len(added)} sheet(s) to {WORKBOOK_PATH.name}:")
    for s in added:
        print(f"  {s}")


if __name__ == "__main__":
    main()
